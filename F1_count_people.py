#!/usr/bin/env python
# coding: utf-8

import numpy as np
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
from openpyxl import workbook 
from openpyxl import load_workbook


def get_confusionmatrix(true_positives,false_positives, false_negatives):
    precision_walkin = true_positives / (true_positives+false_positives)
    recall_walkin = true_positives / (true_positives+false_negatives)
    print(precision_walkin, recall_walkin)
    if precision_walkin and recall_walkin:
        F1_walkin = 2* ((precision_walkin*recall_walkin)/(precision_walkin+recall_walkin))
    else:
        F1_walkin = 0
    return precision_walkin, recall_walkin, F1_walkin


def save_in_excel(wb, mp_name, mp_vdos, gt_person_walkin, mp_person_walkin, precision_walkin, recall_walkin, F1_walkin):
    sheets = wb.sheetnames
    sheet1 = wb[sheets[0]]
    sheet1.cell(row = row, column = 1).value = mp_vdos
    sheet1.cell(row = row, column = 3).value = gt_person_walkin
    sheet1.cell(row = row, column = 4).value = mp_person_walkin
    sheet1.cell(row = row, column = 5).value = precision_walkin
    sheet1.cell(row = row, column = 6).value = recall_walkin 
    sheet1.cell(row = row, column = 7).value = F1_walkin
    wb.save("F1_"+ mp_name +".xlsx")  


#read all columns from groundtruth and machine prediction sheets
def validate_sheet(mp_data, gt_data, row, mp_sheet, mp_path):
    
    mp_name = os.path.basename(mp_path)
    
    #metrics for validation
    true_positives = 0
    false_positives = 0
    false_negatives = 0

    #extract data from mp sheet
    mp_vdos = list(np.unique(mp_data['vdo_name']))
    mp_idx = list(mp_data['idx'])
    mp_person_walkin = np.unique(mp_data['person_walkin'])
    mp_frame = list(mp_data['frame'])

    #extract data from gt sheet 
    gt_vdos = list(np.unique(gt_data['vdo_name']))
    gt_start_frame = list(gt_data['start_frame'])
    gt_end_frame = list(gt_data['end_frame'])
    gt_idx = list(gt_data['idx'])
    gt_person_walkin = gt_data['person_walkin']
    gt_person_walkin.replace('nan', '')
    
    print('Video name:', mp_sheet)
    
    if not mp_data.empty:
        
        for i in range(len(mp_frame)):
            #match idx of mp to check in all idx of gt
            #id idx matches, match the corresponding frame with frame interval in gt

            if mp_idx[i] in gt_idx:
                gt_index = gt_idx.index(mp_idx[i])

                if mp_frame[i] >= gt_start_frame[gt_index] and mp_frame[i] <= gt_end_frame[gt_index]:
                    true_positives += 1

        false_positives = len(mp_frame)-true_positives
        false_negatives = len(gt_start_frame)-true_positives

        precision_walkin, recall_walkin, F1_walkin = get_confusionmatrix(true_positives,false_positives, false_negatives)
        print('---------------------------confusion matrix--------------------------')
        print('precision_walkin:', precision_walkin)
        print('recall_walkin:', recall_walkin)
        print('F1_walkin:', F1_walkin)
        print('\n')
        
        #######write result to excel sheet#############
        
        if not os.path.isfile("F1_"+ mp_name +".xlsx"):
            wb = openpyxl.Workbook()
            save_in_excel(wb, mp_name, str(mp_vdos), int(gt_person_walkin[0]), mp_person_walkin[0], precision_walkin, recall_walkin, F1_walkin)
        else:
            wb = load_workbook("F1_"+ mp_name +".xlsx")
            save_in_excel(wb, mp_name, str(mp_vdos), int(gt_person_walkin[0]), mp_person_walkin[0], precision_walkin, recall_walkin, F1_walkin)
    else:  
        
        wb = load_workbook("F1_"+ mp_name +".xlsx")
        save_in_excel(wb, mp_name, 'N/A', 'N/A', 'N/A', 'N/A', 'N/A', 'N/A')

if __name__ == "__main__":
    
    ap = argparse.ArgumentParser()
    ap.add_argument("--gt", required=True)
    ap.add_argument("--mp", required=True)
 
    args = vars(ap.parse_args())
    gt_path = args['gt']
    mp_path = args['mp']
    
    row = 2
    gt_sheets = pd.ExcelFile(gt_path)
    mp_sheets = pd.ExcelFile(mp_path)
    
    for mp_sheet in mp_sheets.sheet_names:
        mp_data = pd.read_excel(gt_path, sheet_name=mp_sheet)
        if mp_sheet in gt_sheets.sheet_names:
            gt_data = pd.read_excel(mp_path, sheet_name=mp_sheet)
            validate_sheet(mp_data, gt_data, row, mp_sheet, mp_path)
            row+=1
        else:
            print('cant find sheets')





